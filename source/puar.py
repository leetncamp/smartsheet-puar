#!/usr/bin/env python
from __future__ import print_function
print("Starting up...\n")
import os, sys
from pdb import set_trace as debug

BASE_DIR = os.path.abspath(__file__).split("source")[0]
os.chdir(os.path.join(BASE_DIR, "source"))


import traceback
from argparse import ArgumentParser
import re
from jinja2 import Template
import smtplib
import pytz
UTC = pytz.timezone("UTC")
LA = pytz.timezone("America/Los_Angeles")
from snlmailer import Message
from configuration import *  #pulls in variables defined in configuration.py

from collections import OrderedDict
import datetime
import base64
from nameparser import HumanName
import progressbar
import openpyxl

stop_after       = locals().get("stop_after", 0) #stop_after defaults to 0 if not present in configuration.py
default_filename = locals().get("default_filename", "UserAccessReport.xlsx")
smtp_server      = locals().get("smtp_server", "smtp3.hp.com")
log_format       = locals().get("log_format", ".isoformat()")
date_format      = locals().get("date_format", "%Y-%m-%d %H:%M")
format           = locals().get("format", 'WVpFLWFkUi02SlEtcm80')

now = "LA.localize(datetime.datetime.now()){0}".format(log_format)

parser = ArgumentParser(description=u"Notify smartsheet owners of non-HP shares")
parser.add_argument("filename", nargs="?", default=default_filename, help=u"File to read.  Must be XLSX format. Set default in configuration.py")
parser.add_argument("--go", action="store_true", help="with --go, the program does not wait for confirmation. It sends email. Useful for automated running. ")
ns = parser.parse_args()

path2xlsx = os.path.join(BASE_DIR, "smartsheet-puar", ns.filename)

mtime = datetime.datetime.fromtimestamp(os.path.getmtime(path2xlsx))


wb = openpyxl.load_workbook(path2xlsx)
ws =  wb.worksheets[0]


headers = {}

for cell in ws[1]:
    headers[cell.col_idx - 1] = cell.value

count = 0
hpRE = re.compile("@hp.com", re.I)

owners={}

print("Reading spreadsheet written on {0}...\n".format(mtime.strftime("%a %b %d, %Y  %I:%M %p")))

with progressbar.ProgressBar(widgets=[progressbar.Percentage(), progressbar.Bar()], max_value=ws.max_row) as bar:
    for xlrow in ws.iter_rows(row_offset=1):
    	count += 1
        row = {}
        for colnumber, header in headers.iteritems():
            row[header] = xlrow[colnumber].value

        shared_to = row[u"Shared To"]

        if shared_to and not hpRE.search(shared_to):
        	owner = row[u"Owner"]
        	owners[owner] = owners.get(owner, []) + [row]
        	
        else:
        	continue #Skip this row if the Shared-to is an hp address

        bar.update(xlrow[0].row)


if stop_after in locals() and stop_after:
	print(u"\nStopping after {0} emails".format(stop_after))

if not ns.go:
	redirect_emails_to = raw_input(u"\nSending {0} emails. Close the window to cancel. Press Enter to send. Type 'test'\nto redirect all emails to smartsheet.hpadmin@hp.com: ".format(len(owners)))


if not (redirect_emails_to == '' or redirect_emails_to.replace("'","").lower() == 'test' or u"@" in redirect_emails_to):
	print("Exiting. You must either press Return, type 'test', or supply an email address")
	os.system.exit()
if redirect_emails_to == "test":
	redirect_emails_to = "smartsheet.hpadmin@hp.com"



log = file("log.txt", "a")

headers_in_email = [u'Sheet Name', u'Shared To Permission', u'Shared To']
template = Template(open(u"template.html", 'rb').read())
emails_sent = 0
email_log = []

print("Sending emails...")
pb = progressbar.ProgressBar(widgets=[progressbar.Percentage(), progressbar.Bar()], max_value=len(owners)).start()

for owner, rows in owners.iteritems():

	for row in rows:
		#normalize timezone to Pacific Time rather than UTC
		row["Last Modified (PT)"] = LA.normalize(UTC.localize(row["Last Modified Date/Time (UTC)"])).strftime(date_format)
		#row["owner_email"] = owner.replace(u"@hp.com", u"")
		row["Sheet Name"] = row[u"Name"]


	nameStr = u" ".join([item.capitalize() for item in owner.split(u"@")[0].split(".")])
	name = HumanName(nameStr)
	firstname = name.first
	lastname = name.last
	html = template.render(**locals())
	To = redirect_emails_to if redirect_emails_to else owner
	msg = Message(To=To, From=From)
	msg.Subject = Subject
	msg.Html = html
	addlHeaders = [["Precedence","bulk"], ['Disposition-Nofication-To', From]]

	errors = msg.customSend(smtp_server)
	if errors:
		print(errors)
	emails_sent += 1
	dt = eval(now)
	log.write(u"{0}\tEmailed {1}\n".format(dt, msg.To))
	email_log.append(owner)
	pb.update(emails_sent)
	if stop_after and emails_sent >= stop_after:
		print(u"Stopping early at {0} emails. See configuration.py stop_after".format(stop_after))
		break
	

print(u"Sent {0} emails".format(emails_sent))
for owner in email_log:
	print (owner)

if not ns.go:
	raw_input("Press any key to exit.")