#!/usr/bin/env python
from __future__ import print_function

print("Starting...")
"""Convert a spreadsheet into a list of python dictionaries with keys pulled from row 1 as a header.
Converts the first sheet (sheet 0), unless you pass in a sheet number."""

import os
import sys
from pdb import set_trace as debug
import openpyxl
import progressbar

def excel2dict(path2xlsx, sheet=0):

    """supply path2xlsx and get back the first sheet as a list of dictionaries with keys from the (required) header row.
    Converts the first sheet unless you specify sheets=sheetnum"""


    wb = openpyxl.load_workbook(path2xlsx)
    ws =  wb.worksheets[sheet]
    headers = {}

    for cell in ws[1]:
        headers[cell.col_idx - 1] = cell.value

    data = []

    with progressbar.ProgressBar(max_value=ws.max_row) as bar:
        for row in ws.iter_rows(row_offset=1):
            rowdict = {}
            for colnumber, header in headers.iteritems():
                try:
                    rowdict[header] = row[colnumber].value
                except Exception as e:
                    print(traceback.format_exc())
                    raw_input("Press Enter to continue")
            data.append(rowdict)
            bar.update(row[0].row)
    return(data) 


if __name__=="__main__":
    from argparse import ArgumentParser
    parser = ArgumentParser()
    parser.add_argument("path", nargs="?", default="", help="path to excel file in .xlsx format")
    parser.add_argument("--sheet", "-s", default=0, help="sheet to convert")


    ns = parser.parse_args()

    result = excel2dict(ns.path, sheet=ns.sheet)
    print (result)
